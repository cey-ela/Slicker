"""
This module acts as a halfway 'stop over' between the main script and getting to the databases
It prepares the data for 3 destinations: SQL, Excel Ticker Wall and Excel Backup
"""

from kivy.config import Config
from openpyxl import load_workbook
from popups.popups import Popups
from database.sql_connection import update_sql_db

popups = Popups()

kvconfg = Config
kvconfg.read("C:\\Program Files\\Slicker\\slicker.ini")


def convert_categories(category):
    """Before arriving at any of the destinations it goes through this func to suffix CONT. where necessary for the
    Maestro Ticker scene.
    It does so by checking each category against the previous category, if there is a match,  the current cat(egory)
    gets suffixed."""

    first_word = None

    for index, cat in enumerate(category):
        cat = cat.replace("BREAKING", "BREAKING NEWS")
        cat = cat.replace("ENTS", "ENTERTAINMENT")

        if cat != "WEATHER" and cat != "CONTACT US" and cat != "BLANK":
            if cat != first_word:
                first_word = cat
            else:
                cat += " CONT."
        else:
            first_word = None

        category[index] = cat

    return category


def prepare_data_for_sql(data_set_string, data_set):
    """ Then each other function fills in the BLANK categories where needed and adds CONTACT and WEATHER info."""

    try:
        weather_bool = kvconfg.getint('main', 'weather_bool')  # IF weather/contact are to be displayed
        contact_bool = kvconfg.getint('main', 'contact_bool')  # retrieved from the trusty .ini

        index = range(1, 13)
        category = [cat["category"] for cat in data_set]
        text = [t["text"] for t in data_set]

        convert_categories(category)

        for i in range(10 - len(category)):
            category.append("BLANK")

        for i in range(10 - len(text)):
            text.append("")

        # i = index, retrieved using enumerate. Sets the new word at that index position
        # for i, word in enumerate(category):
        #     if word == 'ENTS':
        #         category[i] = 'ENTERTAINMENT'
        #     elif word == 'Ents cont':
        #         category[i] = 'ENTERTAINMENT CONT.'

        if contact_bool:
            category.append("CONTACT US")
            text.append("")
        else:
            category.append("NO CONTACT US")
            text.append("")

        if weather_bool:
            category.append("WEATHER")
            text.append("")
        else:
            category.append("NO WEATHER")
            text.append("")

        data_list_for_sql = [list(a) for a in zip(category, text, index)]  # Zipping lets you loop through 2+ lists
        # at once and join them

        update_sql_db(data_list_for_sql, data_set_string)

    except:
        print('Error sending data to SQL Database')
        popups.sql_warning()  # On screen popup if there are any errors sending the data


def update_excel_wall(data_set_string, data_set):
    """Similar to above but we open a spreadsheet, define which cells we're going to adjust and then write to it"""
    try:
        wall_spreadsheet = "G:\\DataSources\\Ticker\\TickerWall_DB.xlsx"

        weather_bool = kvconfg.getint('main', 'weather_bool')
        contact_bool = kvconfg.getint('main', 'contact_bool')

        index = list(range(1, 10))
        category = [cat["category"] for cat in data_set]
        text = [t["text"] for t in data_set]

        convert_categories(category)

        for i in range(10 - len(category)):
            category.append("Blank")

        for i in range(10 - len(text)):
            text.append("")

        # i = index, retrieved using enumerate. Sets the new word at that index position
        # for i, word in enumerate(category):
        #     if word == 'ENTS':
        #         category[i] = 'ENTERTAINMENT'
        #     elif word == 'Ents Cont':
        #         category[i] = 'ENTERTAINMENT CONT.'

        wb_wall = load_workbook(filename=wall_spreadsheet, read_only=False, keep_vba=True)

        # REFACTOR - if 'generic' in data_set_string

        if 'generic' in data_set_string:
            sheet = wb_wall['TICKER']
        elif 'breaking' in data_set_string:
            sheet = wb_wall["BREAKING NEWS"]
        else:
            sheet = wb_wall['OBIT']

        excel_index = sheet['A2': 'A11']
        excel_cat = sheet['B2':'B11']
        excel_text = sheet['C2':'C11']

        if 'generic' in data_set_string:
            if contact_bool:
                sheet['B13'] = 'CONTACT US'
            else:
                sheet['B13'] = 'NO CONTACT US'

            if weather_bool:
                sheet['B15'] = 'WEATHER'
            else:
                sheet['B15'] = 'NO WEATHER'

        # [item for sublist in excel_cat for item in sublist] is the most efficient way of converting the tuple of
        # tuples, 'wb_wall['TICKER'], into a list of strings that can be populated by the gui_category
        for excel_cell, gui_index in zip([item for sublist in excel_index for item in sublist], index):
            excel_cell.value = gui_index

        for excel_cell, gui_cat in zip([item for sublist in excel_cat for item in sublist], category):
            excel_cell.value = gui_cat

        for excel_cell, gui_text in zip([item for sublist in excel_text for item in sublist], text):
            excel_cell.value = gui_text

        wb_wall.save(wall_spreadsheet)

    except:
        print('Error sending data to Ticker Wall Spreadsheet')
        popups.wall_warning()


# def update_excel_backup(data_set_string, data_set):
#     """Same as the wall data but sending to the backup spreadsheet"""
#     try:
#         backup_spreadsheet = "G:\\DataSources\\Ticker\\Old Files\\TICKER_EDITv4.xlsm"
#
#         weather_bool = kvconfg.getint('main', 'weather_bool')
#         contact_bool = kvconfg.getint('main', 'contact_bool')
#
#         index = list(range(1, 10))
#         category = [cat["category"] for cat in data_set]
#         text = [t["text"] for t in data_set]
#
#         convert_categories(category)
#
#         for i in range(10 - len(category)):
#             category.append("Blank")
#
#         for i in range(10 - len(text)):
#             text.append("")
#
#         # i = index, retrieved using enumerate. Sets the new word at that index position
#         for i, word in enumerate(category):
#             if word == 'ENTS':
#                 category[i] = 'ENTERTAINMENT'
#             elif word == 'Ents Cont':
#                 category[i] = 'ENTERTAINMENT CONT.'
#
#         wb_backup = load_workbook(filename=backup_spreadsheet, read_only=False, keep_vba=True)
#
#         # REFACTOR - if 'generic' in data_set_string
#
#         if 'generic' in data_set_string:
#             sheet = wb_backup['TICKER']
#         elif 'breaking' in data_set_string:
#             sheet = wb_backup["BREAKING NEWS"]
#         else:
#             sheet = wb_backup['OBIT']
#
#         excel_index = sheet['A2': 'A11']
#         excel_cat = sheet['B2':'B11']
#         excel_text = sheet['C2':'C11']
#
#         if 'generic' in data_set_string:
#             if contact_bool:
#                 sheet['B13'] = 'CONTACT US'
#             else:
#                 sheet['B13'] = 'NO CONTACT US'
#
#             if weather_bool:
#                 sheet['B15'] = 'WEATHER'
#             else:
#                 sheet['B15'] = 'NO WEATHER'
#
#         # [item for sublist in excel_cat for item in sublist] is the most efficient way of converting the tuple of
#         # tuples, 'wb_backup['TICKER'], into a list of strings that can be populated by the gui_category
#         for excel_cell, gui_index in zip([item for sublist in excel_index for item in sublist], index):
#             excel_cell.value = gui_index
#
#         for excel_cell, gui_cat in zip([item for sublist in excel_cat for item in sublist], category):
#             excel_cell.value = gui_cat
#
#         for excel_cell, gui_text in zip([item for sublist in excel_text for item in sublist], text):
#             excel_cell.value = gui_text
#
#         wb_backup.save(backup_spreadsheet)
#
#     except:
#         print("Backup excel broken - file probably open somewhere")
#         popups.backup_warning()
